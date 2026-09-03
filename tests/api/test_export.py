from io import BytesIO

from openpyxl import load_workbook


def test_export_returns_xlsx(client, seeded_businesses):
    response = client.get("/export")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    worksheet = load_workbook(BytesIO(response.content)).active
    assert [cell.value for cell in worksheet[1]] == [
        "Nome do Negócio",
        "Tipo de Negócio",
        "Endereço",
        "Telefone",
        "Avaliação",
        "Nº de Avaliações",
        "Possui Site?",
        "Link do Maps",
    ]
    names = {worksheet[f"A{row}"].value for row in range(2, worksheet.max_row + 1)}
    assert names == {"Loja Com Site", "Padaria Sem Site"}


def test_export_only_without_website(client, seeded_businesses):
    response = client.get("/export", params={"only_without_website": True})
    assert response.status_code == 200
    worksheet = load_workbook(BytesIO(response.content)).active
    names = {worksheet[f"A{row}"].value for row in range(2, worksheet.max_row + 1)}
    assert names == {"Padaria Sem Site"}


def test_export_fills_every_column_with_the_business_own_data(client, seeded_businesses):
    with_site, without_site = seeded_businesses
    response = client.get("/export")
    assert response.status_code == 200

    worksheet = load_workbook(BytesIO(response.content)).active
    rows_by_name = {
        row[0].value: [cell.value for cell in row]
        for row in worksheet.iter_rows(min_row=2)
    }

    assert rows_by_name[with_site.name] == [
        with_site.name,
        "Loja de Roupas",
        with_site.address,
        with_site.phone,
        None,
        None,
        "Sim",
        with_site.maps_url,
    ]
    assert rows_by_name[without_site.name] == [
        without_site.name,
        "Padaria",
        without_site.address,
        without_site.phone,
        None,
        None,
        "Não",
        without_site.maps_url,
    ]


def test_export_only_the_selected_search_results(client, db_session, seeded_businesses):
    with_site, without_site = seeded_businesses
    extra_business = type(with_site)(
        place_id="place-extra",
        name="Lead Fora da Busca",
        has_website=False,
        business_type="bakery",
    )
    db_session.add(extra_business)
    db_session.commit()

    response = client.get("/export", params=[("ids", without_site.id), ("ids", with_site.id)])
    worksheet = load_workbook(BytesIO(response.content)).active
    names = {worksheet[f"A{row}"].value for row in range(2, worksheet.max_row + 1)}

    assert names == {"Loja Com Site", "Padaria Sem Site"}


def test_export_applies_history_filters(client, seeded_businesses):
    response = client.get(
        "/export",
        params={"name": "Padaria", "has_website": "false", "business_type": "bakery"},
    )
    worksheet = load_workbook(BytesIO(response.content)).active
    names = {worksheet[f"A{row}"].value for row in range(2, worksheet.max_row + 1)}

    assert names == {"Padaria Sem Site"}
