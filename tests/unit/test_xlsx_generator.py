from types import SimpleNamespace

from openpyxl import load_workbook

from services.business_types import get_business_type
from services.xlsx_generator import generate_xlsx


def test_xlsx_headers_and_business_type_label():
    category = get_business_type("caca_vazamento")
    workbook = generate_xlsx(
        [
            SimpleNamespace(
                name="Exemplo",
                business_type="caca_vazamento",
                address="Rua A",
                phone="123",
                has_website=False,
                maps_url="",
            )
        ]
    )
    worksheet = load_workbook(workbook).active
    assert [cell.value for cell in worksheet[1]] == [
        "Nome do Negócio",
        "Tipo de Negócio",
        "Endereço",
        "Telefone",
        "Avaliação",
        "Nº de Avaliações",
        "Possui Site?",
        "Link do Maps",
    ]
    assert [cell.value for cell in worksheet[2]] == [
        "Exemplo",
        category.label,
        "Rua A",
        "123",
        None,
        None,
        "Não",
        None,
    ]


def test_xlsx_writes_every_column_for_each_row_without_shifting():
    # Duas linhas com valores bem diferentes entre si — pega qualquer bug de
    # coluna/linha trocada (ex.: rating indo pra célula errada, ou a segunda
    # linha herdando valor da primeira).
    category = get_business_type("caca_vazamento")
    rows = generate_xlsx(
        [
            SimpleNamespace(
                name="Padaria do João",
                business_type="caca_vazamento",
                address="Rua das Flores, 123",
                phone="+55 48 99999-1234",
                has_website=False,
                maps_url="https://maps.google.com/?cid=1",
                rating=4.5,
                user_ratings_total=120,
            ),
            SimpleNamespace(
                name="Oficina Mecânica Silva",
                business_type=None,
                address=None,
                phone=None,
                has_website=True,
                maps_url=None,
                rating=None,
                user_ratings_total=None,
            ),
        ]
    )
    ws = load_workbook(rows).active
    assert ws.max_row == 3  # cabeçalho + 2 negócios

    assert [c.value for c in ws[2]] == [
        "Padaria do João",
        category.label,
        "Rua das Flores, 123",
        "+55 48 99999-1234",
        4.5,
        120,
        "Não",
        "https://maps.google.com/?cid=1",
    ]

    # Segunda linha exercita os fallbacks: sem endereço/telefone/tipo/maps_url/nota.
    assert [c.value for c in ws[3]] == [
        "Oficina Mecânica Silva",
        "Não informado",
        "Não informado",
        "Não informado",
        None,
        None,
        "Sim",
        None,
    ]
