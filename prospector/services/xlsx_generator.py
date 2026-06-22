from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from services.places import CATEGORY_CONFIG  # type: ignore


def generate_xlsx(businesses: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Negócios Prospectados"

    headers = ["Nome do Negócio", "Endereço", "Telefone", "Tipo", "Possui Site?", "Link do Maps"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row, biz in enumerate(businesses, 2):
        biz_t = biz.business_type or ""
        cat_conf = CATEGORY_CONFIG.get(biz_t, {})
        # If the category config has queries, use the first query as the display name, else use the raw type
        queries = cat_conf.get("queries", [biz_t])
        display_type = str(queries[0]).title() if queries else "Não informado"
        
        ws.cell(row=row, column=1, value=biz.name)
        ws.cell(row=row, column=2, value=biz.address or "Não informado")
        ws.cell(row=row, column=3, value=biz.phone or "Não informado")
        ws.cell(row=row, column=4, value=display_type)
        ws.cell(row=row, column=5, value="Sim" if biz.has_website else "Não")
        ws.cell(row=row, column=6, value=biz.maps_url or "")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 55

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
