from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from services.business_types import type_labels


def generate_xlsx(businesses: list) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Negócios Prospectados"

    headers = [
        "Nome do Negócio",
        "Tipo de Negócio",
        "Endereço",
        "Telefone",
        "Avaliação",
        "Nº de Avaliações",
        "Possui Site?",
        "Link do Maps",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    labels = type_labels()
    for row, biz in enumerate(businesses, 2):
        ws.cell(row=row, column=1, value=biz.name)
        ws.cell(row=row, column=2, value=labels.get(biz.business_type, biz.business_type or "Não informado"))
        ws.cell(row=row, column=3, value=biz.address or "Não informado")
        ws.cell(row=row, column=4, value=biz.phone or "Não informado")
        rating = getattr(biz, "rating", None)
        ratings_total = getattr(biz, "user_ratings_total", None)
        ws.cell(row=row, column=5, value=rating if rating is not None else "")
        ws.cell(row=row, column=6, value=ratings_total if ratings_total is not None else "")
        ws.cell(row=row, column=7, value="Sim" if biz.has_website else "Não")
        ws.cell(row=row, column=8, value=biz.maps_url or "")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 55

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
